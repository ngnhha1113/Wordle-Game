# astar_wordle.py
import heapq
import time
from collections import Counter
import matplotlib.pyplot as plt
import random
from openpyxl import Workbook, load_workbook
import os
# -----------------------------
# Example API containing target + dictionary
# -----------------------------


# -----------------------------
# A* Solver Class
# -----------------------------
class AStarSolver:
    def __init__(self, api):
        self.api = api
        self.target = api.word
        self.words = api.words_list
        self.N = len(self.target)

        # Measurement
        self.expanded_nodes = []
        self.max_open_size = 0
        self.max_memory = 0
        self.search_time = 0
        self.expanded_nodes_list = []
    # ------------------------ Heuristic ------------------------
    def heuristic(self, w):
    
        mismatch = 0
        for i in range(self.N):
            if w[i] != self.target[i]:
                mismatch += 1  # chữ sai vị trí → penalty

    # Count letters
        wc = Counter(w)
        tc = Counter(self.target)

    # Penalty cho chữ không thuộc target
        extra = sum(wc[ch] for ch in wc if ch not in tc)

    # Penalty chênh lệch tần suất
    # (vd target có 2 chữ L nhưng word có 0)
        freq_mismatch = 0
        for ch in set(wc.keys()).union(tc.keys()):
            freq_mismatch += abs(wc[ch] - tc[ch])

        # Weighting
        return mismatch * 2 + extra * 3 + freq_mismatch


    # ------------------------ Feedback ------------------------
    def feedback(self, guess, target):
        res = ["B"] * self.N
        t_left = list(target)

        # Green letters
        for i in range(self.N):
            if guess[i] == target[i]:
                res[i] = "G"
                t_left[i] = None

        # Yellow letters
        for i in range(self.N):
            if res[i] == "G":
                continue
            if guess[i] in t_left:
                res[i] = "Y"
                t_left[t_left.index(guess[i])] = None

        return "".join(res)

    # ------------------------ Consistency Check ------------------------
    def consistent(self, word, history):
        for g, fb in history:
            if self.feedback(word, g) != fb:
                return False
        return True

    # ------------------------ A* Search ------------------------
    def solve(self, start="AARON"):
        
        start_time = time.time()
        pq = []
        h = []
        heapq.heappush(pq, (0, start, []))
        visited = set()
        self.expanded_nodes_list = []
        while pq:
            self.max_open_size = max(self.max_open_size, len(pq))
            f, guess, history = heapq.heappop(pq)

            if guess in visited:
                continue
            visited.add(guess)
            
            self.max_memory = max(self.max_memory, len(visited) + len(pq))
            self.expanded_nodes_list.append(guess)
            self.expanded_nodes.append(guess)
            fb = self.feedback(guess, self.target)
            new_history = history + [(guess, fb)]
            
            if guess == self.target:
                self.search_time = time.time() - start_time
                return [g for g, _ in new_history]

            for w in self.words:
                if w in visited:
                    continue
                if not self.consistent(w, new_history):
                    continue

                g_cost = len(new_history)
                h_cost = self.heuristic(w)
                heapq.heappush(pq, (g_cost + h_cost, w, new_history))

        self.search_time = time.time() - start_time
        return []

    # ------------------------ Export Statistics ------------------------
    def get_stats(self):
        print("Expanded Nodes:")
        for node in self.expanded_nodes_list:
            print(node)
        avg_guess = len(self.expanded_nodes_list)  # số guesses trung bình
        return {
            "target": self.target,
            "expanded_nodes": len(self.expanded_nodes_list),
            "max_open_size": self.max_open_size,
            "max_memory_nodes": self.max_memory,
            "search_time_sec": round(self.search_time, 4),
            
        }

    # ------------------------ Plot performance ------------------------
    def plot_stats(self):
        stats = self.get_stats()
        plt.bar(['Expanded Nodes', 'Max Open Size', 'Max Memory'],
                [stats['expanded_nodes'], stats['max_open_size'], stats['max_memory_nodes']])
        plt.title(f"A* Wordle Search Performance (Target: {stats['target']})")
        plt.show()
    #def print_expanded_nodes(self):
    def save_to_excel(self, stats):
        filename = "Experiments.xlsx"
    
    # Nếu file chưa tồn tại → tạo mới
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "AStar"

            # Header
            ws.append(["Algorithm", "time", "memory", "expanded_nodes",
                       "target", "list_expanded_nodes"])
        else:
            wb = load_workbook(filename)
            ws = wb.active

    # Ghi dữ liệu
        ws.append([
            "A*",                             # Algorithm
            stats["search_time_sec"],         # time
            stats["max_memory_nodes"],        # memory
            stats["expanded_nodes"],          # expanded_nodes
            stats["target"],                  # target word
            ", ".join(self.expanded_nodes_list)   # list expanded node
        ])

        wb.save(filename)       

